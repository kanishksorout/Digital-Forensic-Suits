import pandas as pd
import browser_history
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def fetch_and_save_history(browser_name, workbook):
    browser = None
    if browser_name == 'Firefox':
        from browser_history.browsers import Firefox
        browser = Firefox()
    elif browser_name == 'Chrome':
        from browser_history.browsers import Chrome
        browser = Chrome()
    elif browser_name == 'Brave':
        from browser_history.browsers import Brave
        browser = Brave()
    elif browser_name == 'Opera':
        from browser_history.browsers import Opera
        browser = Opera()
    elif browser_name == 'Chromium':
        from browser_history.browsers import Chromium
        browser = Chromium()
    elif browser_name == 'Edge':
        from browser_history.browsers import Edge
        browser = Edge()
    elif browser_name == 'LibreWolf':
        from browser_history.browsers import LibreWolf
        browser = LibreWolf()
    elif browser_name == 'OperaGX':
        from browser_history.browsers import OperaGX
        browser = OperaGX()
    elif browser_name == 'Vivaldi':
        from browser_history.browsers import Vivaldi
        browser = Vivaldi()

    if browser is not None:
        outputs = browser.fetch_history()
        history_data = outputs.histories
        df = pd.DataFrame(history_data, columns=['DateTime', 'URL'])
        df['DateTime'] = df['DateTime'].apply(lambda dt: dt.replace(tzinfo=None))
        df.sort_values(by='DateTime', inplace=True)
        append_to_excel(workbook, browser_name, df)

def append_to_excel(workbook, sheet_name, df):
    try:
        sheet = workbook[sheet_name]
    except KeyError:
        sheet = workbook.create_sheet(title=sheet_name)

    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

def main():
    browsers = ['Firefox', 'Chrome', 'Brave', 'Opera', 'Chromium', 'Edge', 'LibreWolf', 'OperaGX', 'Vivaldi']
    current_directory = os.getcwd()
    excel_filename = os.path.join(current_directory, 'browser_history.xlsx')

    # Cleanup existing file if it exists
    if os.path.exists(excel_filename):
        os.remove(excel_filename)

    workbook = Workbook()

    for browser_name in browsers:
        fetch_and_save_history(browser_name, workbook)

    workbook.save(excel_filename)

if __name__ == "__main__":
    main()